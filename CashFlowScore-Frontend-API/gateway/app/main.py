import base64
import csv
import io
import json
import os
import requests
from typing import Any, Dict, List
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

from fastapi import FastAPI, File, HTTPException, UploadFile, Depends, Security
from fastapi.security import APIKeyHeader
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from pydantic import BaseModel

from .mock_data import get_mock_businesses, get_mock_status

# ─── Service URLs ─────────────────────────────────────────────────────────────
PIPELINE_URL = os.getenv("PAVAN_PIPELINE_URL", "http://127.0.0.1:8002")

app = FastAPI(title="CashFlowScore Gateway", version="0.1.0")

API_KEY = "NBFC_TEST_KEY_123"
api_key_header = APIKeyHeader(name="Authorization", auto_error=False)

async def verify_api_key(api_key: str = Security(api_key_header)):
    if not api_key or not api_key.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Missing or invalid Authorization header")
    token = api_key.replace("Bearer ", "")
    if token != API_KEY:
        raise HTTPException(status_code=403, detail="Invalid API Key")
    return token

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://127.0.0.1:5173",
        "http://localhost:5173",
        "http://127.0.0.1:3000",
        "http://localhost:3000",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


class ScoreRequest(BaseModel):
    business_id: str | None = None
    features: Dict[str, Any] | None = None


def _calculate_score(features: Dict[str, Any]) -> int:
    inflow = float(features.get("inflow_amount", 0) or 0)
    gst_delay = float(features.get("gst_delay_days", 0) or 0)
    bounce_count = float(features.get("bounce_count", 0) or 0)

    score = 42
    if inflow >= 350000:
        score += 24
    elif inflow >= 220000:
        score += 14
    elif inflow >= 100000:
        score += 6

    if gst_delay <= 2:
        score += 18
    elif gst_delay <= 5:
        score += 8

    if bounce_count == 0:
        score += 15
    elif bounce_count == 1:
        score += 6

    score -= min(20, bounce_count * 5 + max(0, gst_delay - 2) * 3)
    return max(0, min(100, int(score)))


def _build_reasons(features: Dict[str, Any]) -> List[str]:
    inflow = float(features.get("inflow_amount", 0) or 0)
    gst_delay = float(features.get("gst_delay_days", 0) or 0)
    bounce_count = float(features.get("bounce_count", 0) or 0)

    reasons: List[str] = []
    if inflow >= 300000:
        reasons.append("Cash inflow is strong and recurring")
    elif inflow >= 180000:
        reasons.append("Cash inflow is moderate and improving")
    else:
        reasons.append("Cash inflow is below the target band")

    if gst_delay <= 2:
        reasons.append("GST filing cadence is healthy")
    elif gst_delay <= 5:
        reasons.append("GST filing is slightly delayed")
    else:
        reasons.append("GST filing delays are a material risk")

    if bounce_count == 0:
        reasons.append("Bounce frequency remains low")
    elif bounce_count == 1:
        reasons.append("Bounce frequency is manageable")
    else:
        reasons.append("Bounce frequency is elevated")

    return reasons[:3]


def _get_json(url: str) -> Dict[str, Any]:
    request = Request(url, headers={"Accept": "application/json"}, method="GET")
    with urlopen(request, timeout=10) as response:
        return json.loads(response.read().decode("utf-8"))


def _post_json(url: str, payload: Dict[str, Any]) -> Dict[str, Any]:
    data = json.dumps(payload).encode("utf-8")
    request = Request(
        url,
        data=data,
        headers={"Content-Type": "application/json", "Accept": "application/json"},
        method="POST",
    )
    with urlopen(request, timeout=10) as response:
        return json.loads(response.read().decode("utf-8"))


def _load_rows_from_upload(content: bytes, filename: str) -> List[Dict[str, Any]]:
    name = (filename or "").lower()
    if name.endswith(".csv"):
        text = content.decode("utf-8")
        return list(csv.DictReader(io.StringIO(text)))

    if name.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise HTTPException(status_code=400, detail="openpyxl is required for .xlsx uploads") from exc

        workbook = load_workbook(io.BytesIO(content), data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(cell or "").strip() for cell in rows[0]]
        parsed_rows: List[Dict[str, Any]] = []
        for row in rows[1:]:
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue
            parsed_row = {}
            for index, header in enumerate(headers):
                value = row[index] if index < len(row) else None
                parsed_row[header] = value
            parsed_rows.append(parsed_row)
        return parsed_rows

    raise HTTPException(status_code=400, detail="expected a .csv or .xlsx file")


@app.get("/health")
def health() -> Dict[str, str]:
    return {"status": "ok", "service": "gateway"}


def _map_pipeline_business(item: Dict[str, Any]) -> Dict[str, Any]:
    biz_id = item["business_id"]
    profile = item.get("profile") or {}
    features = item.get("features") or {}
    
    # Calculate score & reasons via upstream ML service or fallback local logic
    score = 50
    reasons = ["Score calculated dynamically"]
    
    upstream_url = os.getenv("AKAASH_SCORE_URL")
    if upstream_url:
        try:
            upstream_result = _post_json(upstream_url, {"business_id": biz_id, "features": features})
            if isinstance(upstream_result, dict):
                score = int(upstream_result.get("score", 50))
                reasons = upstream_result.get("top_reasons") or upstream_result.get("reasons") or ["Calculation succeeded"]
        except Exception:
            pass
            
    if not reasons or reasons == ["Score calculated dynamically"]:
        # Fallback to local heuristic score
        inflow = float(features.get("monthly_upi_volume", 250000))
        gst_reg = float(features.get("gst_filing_regularity", 90))
        bounces = float(features.get("bounce_frequency", 0))
        
        # Mapping inputs for heuristic calculation
        calc_features = {
            "inflow_amount": inflow,
            "gst_delay_days": max(0, int(20 - (gst_reg / 5))),
            "bounce_count": bounces
        }
        score = _calculate_score(calc_features)
        reasons = _build_reasons(calc_features)

    # Name mapping
    name = profile.get("business_name") or profile.get("name")
    if not name:
        name = f"MSME {biz_id[-4:] if len(biz_id) >= 4 else biz_id}".upper()
        
    segment = profile.get("segment") or "Retail"
    
    return {
        "id": biz_id,
        "name": name,
        "segment": segment,
        "score": score,
        "status": "approved" if score >= 70 else "rejected",
        "credit_unlocked": int(float(features.get("monthly_upi_volume", 250000)) * 1.5) if score >= 70 else 0,
        "risk_band": "low" if score >= 80 else "medium" if score >= 60 else "high",
        "recent_activity": f"Ingestion active via {features.get('source', 'stream')}" if features else "Created in pipeline",
        "reasons": reasons,
        "editable_inputs": {
            "business_age_years": int(float(features.get("business_age_years", 5))),
            "monthly_upi_volume": int(float(features.get("monthly_upi_volume", features.get("inflow_amount", 250000)))),
            "monthly_bank_volume": int(float(features.get("monthly_bank_volume", 50000))),
            "monthly_cash_volume": int(float(features.get("monthly_cash_volume", 10000))),
            "gst_filing_regularity": int(float(features.get("gst_filing_regularity", max(0, 100 - float(features.get("gst_delay_days", 2)) * 5)))),
            "gst_turnover": int(float(features.get("gst_turnover", 500000))),
            "bounce_frequency": int(float(features.get("bounce_frequency", features.get("bounce_count", 0)))),
            "avg_monthly_balance": int(float(features.get("avg_monthly_balance", 100000))),
            "income_stability": float(features.get("income_stability", 0.8)),
            "seasonality_score": float(features.get("seasonality_score", 0.5)),
            "loan_default_history": int(float(features.get("loan_default_history", 0)))
        }
    }


@app.get("/businesses")
def get_businesses() -> List[Dict[str, Any]]:
    status_url = os.getenv("PAVAN_STATUS_URL")
    if status_url:
        try:
            base_url = status_url.rsplit("/pipeline-status", 1)[0]
            pipeline_data = _get_json(f"{base_url}/businesses")
            items = pipeline_data.get("items") or []
            if items:
                return [_map_pipeline_business(item) for item in items]
        except Exception as e:
            print(f"Failed to fetch businesses from pipeline, falling back to heuristic: {e}")
            
    return get_mock_businesses()


@app.get("/businesses/{business_id}")
def get_business_detail(business_id: str) -> Dict[str, Any]:
    status_url = os.getenv("PAVAN_STATUS_URL")
    if status_url:
        try:
            base_url = status_url.rsplit("/pipeline-status", 1)[0]
            item = _get_json(f"{base_url}/businesses/{business_id}")
            if item:
                return _map_pipeline_business(item)
        except Exception:
            pass
            
    for business in get_mock_businesses():
        if business["id"] == business_id:
            return business
    raise HTTPException(status_code=404, detail="business not found")


@app.post("/score")
def score_business(payload: ScoreRequest, api_key: str = Depends(verify_api_key)) -> Dict[str, Any]:
    business = None
    for item in get_mock_businesses():
        if item["id"] == payload.business_id:
            business = item
            break

    if business is None and not payload.business_id:
        raise HTTPException(status_code=400, detail="business_id or features are required")

    features = dict(payload.features or {})
    if business and not features:
        features = dict(business.get("editable_inputs", {}))

    upstream_url = os.getenv("AKAASH_SCORE_URL")
    if upstream_url:
        ml_features = {
            "business_age_years": float(features.get("business_age_years", 5)),
            "monthly_upi_volume": float(features.get("monthly_upi_volume", features.get("inflow_amount", 250000))),
            "monthly_bank_volume": float(features.get("monthly_bank_volume", 50000)),
            "monthly_cash_volume": float(features.get("monthly_cash_volume", 10000)),
            "gst_filing_regularity": float(features.get("gst_filing_regularity", max(0, 100 - float(features.get("gst_delay_days", 2)) * 5))),
            "gst_turnover": float(features.get("gst_turnover", 500000)),
            "bounce_frequency": float(features.get("bounce_frequency", features.get("bounce_count", 0))),
            "avg_monthly_balance": float(features.get("avg_monthly_balance", 100000)),
            "income_stability": float(features.get("income_stability", 0.8)),
            "seasonality_score": float(features.get("seasonality_score", 0.5)),
            "loan_default_history": float(features.get("loan_default_history", 0)),
        }
        try:
            upstream_result = _post_json(upstream_url, {"business_id": payload.business_id, "features": ml_features})
            if isinstance(upstream_result, dict):
                return {
                    "score": int(upstream_result.get("score", _calculate_score(features))),
                    "top_reasons": upstream_result.get("top_reasons") or upstream_result.get("reasons") or _build_reasons(features),
                    "source": "xgboost",
                    "business_id": payload.business_id,
                }
        except (HTTPError, URLError, TimeoutError, ValueError, json.JSONDecodeError):
            pass

    score = _calculate_score(features)
    reasons = _build_reasons(features)

    return {
        "score": score,
        "top_reasons": reasons,
        "source": "heuristic",
        "business_id": payload.business_id,
    }


@app.post("/score-batch")
async def score_batch(file: UploadFile = File(...), api_key: str = Depends(verify_api_key)) -> JSONResponse:
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="empty file")

    upstream_url = os.getenv("AKAASH_SCORE_URL")
    if upstream_url:
        try:
            # Determine the batch endpoint url
            batch_url = upstream_url.replace("/score", "/score-batch") if upstream_url.endswith("/score") else f"{upstream_url}/score-batch"
            
            # Forward the file to the ML service
            files = {"file": (file.filename, content, file.content_type)}
            response = requests.post(batch_url, files=files, timeout=30)
            
            if response.status_code == 200:
                # The ML service might return a file directly (as per its FileResponse)
                # We need to package it for the frontend
                result_content = response.content
                
                # Parse the returned Excel to generate summary stats for the UI
                rows = _load_rows_from_upload(result_content, "scored_output.xlsx")
                
                # Create download payload
                download_filename = "scored_output.xlsx"
                download_content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                download_payload = base64.b64encode(result_content).decode("utf-8")
                
                summary = {
                    "rows_processed": len(rows),
                    "preview_count": min(50, len(rows)),
                    "approved_count": sum(1 for row in rows if float(row.get("credit_score", 0)) >= 70),
                    "rejected_count": sum(1 for row in rows if float(row.get("credit_score", 0)) < 70),
                    "average_score": round(sum(float(row.get("credit_score", 0)) for row in rows) / len(rows), 1) if rows else 0,
                }
                
                preview_rows = rows[:50]
                # Format preview rows to match UI expectations
                formatted_preview = []
                for row in preview_rows:
                    formatted_preview.append({
                        "business_name": row.get("business_name", "Unknown"),
                        "inflow_amount": row.get("inflow_amount", 0),
                        "gst_delay_days": row.get("gst_delay_days", 0),
                        "bounce_count": row.get("bounce_count", 0),
                        "score": float(row.get("credit_score", 0)),
                        "top_reasons": row.get("top_reasons", ""),
                    })

                return JSONResponse(
                    status_code=200,
                    content={
                        "message": "batch scoring completed via ML service",
                        "rows_processed": len(rows),
                        "download_filename": download_filename,
                        "download_content_type": download_content_type,
                        "download_content_base64": download_payload,
                        "preview": formatted_preview,
                        "summary": summary,
                    },
                )
        except Exception as e:
            print(f"Upstream ML service failed for batch scoring, falling back to heuristic: {e}")

    # Fallback to local logic
    rows = _load_rows_from_upload(content, file.filename or "")
    if not rows:
        raise HTTPException(status_code=400, detail="no rows found in uploaded file")

    scored_rows: List[Dict[str, Any]] = []
    for row in rows:
        features = {
            "inflow_amount": float(row.get("inflow_amount", 0) or 0),
            "gst_delay_days": float(row.get("gst_delay_days", 0) or 0),
            "bounce_count": float(row.get("bounce_count", 0) or 0),
        }
        score = _calculate_score(features)
        scored_rows.append(
            {
                "business_name": row.get("business_name", "Unknown"),
                "inflow_amount": features["inflow_amount"],
                "gst_delay_days": features["gst_delay_days"],
                "bounce_count": features["bounce_count"],
                "score": score,
                "top_reasons": " | ".join(_build_reasons(features)),
            }
        )

    # ── Publish scored rows as real events to the pipeline ──────────────────
    import uuid
    from datetime import datetime, timezone
    try:
        events_to_publish = []
        for i, row in enumerate(scored_rows):
            biz_id = f"UPLOAD-{abs(hash(row['business_name'])) % 100000:05d}"
            events_to_publish.append({
                "event_id": str(uuid.uuid4()),
                "business_id": biz_id,
                "topic": "txn.upi",
                "event_type": "bounce" if row["bounce_count"] > 0 else "credit",
                "amount": row["inflow_amount"],
                "direction": "inflow",
                "balance_after": row["inflow_amount"] * 1.2,
                "timestamp": datetime.now(timezone.utc).isoformat(),
                "metadata": {"source": "batch_upload", "score": row["score"]},
            })
        if events_to_publish:
            requests.post(
                f"{PIPELINE_URL}/transactions/bulk",
                json=events_to_publish,
                timeout=10,
            )
    except Exception as _pub_err:
        pass  # Non-blocking — scoring still works if pipeline is down

    # Always output a clean .xlsx report for the NBFC (professional, opens in Excel)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, numbers
    except ImportError as exc:
        raise HTTPException(status_code=400, detail="openpyxl is required") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "CashFlowScore Report"

    # Header row with formatting
    headers = ["Business Name", "Monthly Inflow (₹)", "GST Delay (days)", "Bounces", "Credit Score", "Status", "Risk Band", "Credit Unlocked (₹)", "Top Reasons"]
    sheet.append(headers)

    # Style header row
    header_fill = PatternFill(start_color="0D3D4A", end_color="0D3D4A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    green_fill  = PatternFill(start_color="0A2E1A", end_color="0A2E1A", fill_type="solid")
    red_fill    = PatternFill(start_color="2E0A0A", end_color="2E0A0A", fill_type="solid")
    amber_fill  = PatternFill(start_color="2E1E0A", end_color="2E1E0A", fill_type="solid")

    for row in scored_rows:
        sc = row["score"]
        status_label = "Approved" if sc >= 70 else "Rejected"
        risk_band    = "Low" if sc >= 80 else "Medium" if sc >= 60 else "High"
        credit_unlocked = int(row["inflow_amount"] * 1.5) if sc >= 70 else 0
        sheet.append([
            row["business_name"],
            int(row["inflow_amount"]),
            int(row["gst_delay_days"]),
            int(row["bounce_count"]),
            sc,
            status_label,
            risk_band,
            credit_unlocked,
            row["top_reasons"],
        ])
        # Color the score cell
        score_cell = sheet.cell(sheet.max_row, 5)
        if sc >= 75:
            score_cell.fill = green_fill
            score_cell.font = Font(bold=True, color="10B981")
        elif sc >= 55:
            score_cell.fill = amber_fill
            score_cell.font = Font(bold=True, color="F59E0B")
        else:
            score_cell.fill = red_fill
            score_cell.font = Font(bold=True, color="F43F5E")

    # Auto-width columns
    for col in sheet.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        sheet.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    workbook_buffer = io.BytesIO()
    workbook.save(workbook_buffer)
    workbook_bytes = workbook_buffer.getvalue()
    download_filename = "CashFlowScore_Report.xlsx"
    download_content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    download_payload = base64.b64encode(workbook_bytes).decode("utf-8")

    summary = {
        "rows_processed": len(scored_rows),
        "preview_count": min(50, len(scored_rows)),
        "approved_count": sum(1 for row in scored_rows if row["score"] >= 70),
        "rejected_count": sum(1 for row in scored_rows if row["score"] < 70),
        "average_score": round(sum(row["score"] for row in scored_rows) / len(scored_rows), 1) if scored_rows else 0,
    }

    return JSONResponse(
        status_code=200,
        content={
            "message": "batch scoring completed",
            "rows_processed": len(scored_rows),
            "download_filename": download_filename,
            "download_content_type": download_content_type,
            "download_content_base64": download_payload,
            "preview": scored_rows[:50],
            "summary": summary,
        },
    )


@app.get("/status")
def get_status() -> Dict[str, Any]:
    # Try real pipeline first
    try:
        pipeline_status = _get_json(f"{PIPELINE_URL}/pipeline-status")
        if isinstance(pipeline_status, dict):
            return pipeline_status
    except Exception:
        pass
    # Legacy env var fallback
    upstream_url = os.getenv("PAVAN_STATUS_URL")
    if upstream_url:
        try:
            upstream_status = _get_json(upstream_url)
            if isinstance(upstream_status, dict):
                return upstream_status
        except (HTTPError, URLError, TimeoutError, ValueError, json.JSONDecodeError):
            pass
    return get_mock_status()


@app.get("/activity-feed")
def get_activity_feed(limit: int = 30) -> Dict[str, Any]:
    """Proxy to the real event pipeline activity feed — shows live Kafka events."""
    try:
        data = _get_json(f"{PIPELINE_URL}/activity-feed?limit={limit}")
        return data
    except Exception:
        return get_mock_businesses()


@app.get("/pipeline-stats")
def get_pipeline_stats_live() -> Dict[str, Any]:
    """Returns live pipeline stats: queue depth, event counts, business count."""
    try:
        data = _get_json(f"{PIPELINE_URL}/pipeline-status")
        return {
            "queue_depth": data.get("queue_depth", 0),
            "events_processed": data.get("events_processed", 0),
            "business_count": data.get("business_count", 0),
            "stress_flag_count": data.get("stress_flag_count", 0),
            "cache_hit_rate": data.get("cache_hit_rate", 0),
            "redpanda_up": data.get("redpanda_up", False),
            "redis_up": data.get("redis_up", False),
            "postgres_up": data.get("postgres_up", False),
            "last_event_at": data.get("last_event_at"),
            "event_count_by_topic": data.get("event_count_by_topic", {}),
            "timescaledb_rows": data.get("timescaledb_rows"),
            "redis_keys": data.get("redis_keys"),
            "dedup_backend": data.get("dedup_backend", "memory"),
        }
    except Exception:
        return {"queue_depth": 0, "events_processed": 0, "business_count": 0}


# ── Data Cleaning proxy ────────────────────────────────────────────────────────

@app.get("/data-cleaning/raw")
def proxy_data_cleaning_raw() -> Dict[str, Any]:
    try:
        return _get_json(f"{PIPELINE_URL}/data-cleaning/raw")
    except Exception as e:
        raise HTTPException(status_code=502, detail=str(e))


@app.post("/data-cleaning/clean")
def proxy_data_cleaning_clean(api_key: str = Depends(verify_api_key)) -> Dict[str, Any]:
    try:
        return _post_json(f"{PIPELINE_URL}/data-cleaning/clean", {})
    except Exception as e:
        raise HTTPException(status_code=502, detail=str(e))


@app.get("/pipeline")
def get_pipeline_stats():
    """Returns live stats from every component in the system design pipeline."""
    result = {}

    try:
        import redis as redis_lib
        rc = redis_lib.Redis(host="localhost", port=6379, socket_connect_timeout=3, socket_timeout=3)
        rc.ping()
        info = rc.info()
        result["redis"] = {
            "status": "up",
            "total_keys": rc.dbsize(),
            "hits": info.get("keyspace_hits", 0),
            "misses": info.get("keyspace_misses", 0),
            "memory_mb": round(info.get("used_memory", 0) / 1024 / 1024, 2),
            "ops_per_sec": info.get("instantaneous_ops_per_sec", 0),
        }
    except Exception as e:
        result["redis"] = {"status": "down", "error": str(e)[:80]}

    try:
        import psycopg2
        conn = psycopg2.connect(host="localhost", port=5432, dbname="cashflowscore", user="cashflow", password="cashflow", connect_timeout=5)
        cur = conn.cursor()
        cur.execute("SELECT count(*) FROM information_schema.tables WHERE table_schema = 'public';")
        table_count = cur.fetchone()[0]
        row_count = 0
        try:
            # Table is named 'events' (matches store.py schema + docker/initdb/01_schema.sql)
            cur.execute("SELECT count(*) FROM events;")
            row_count = cur.fetchone()[0]
        except Exception:
            pass
        conn.close()
        result["timescaledb"] = {"status": "up", "tables": table_count, "event_rows": row_count}
    except Exception as e:
        result["timescaledb"] = {"status": "down", "error": str(e)[:120]}

    try:
        import requests as req
        rp = req.get("http://localhost:9644/v1/cluster/health_overview", timeout=4)
        if rp.status_code == 200:
            health = rp.json()
            brokers_resp = req.get("http://localhost:9644/v1/brokers", timeout=4)
            broker_count = len(brokers_resp.json()) if brokers_resp.status_code == 200 else 1
            topics_resp = req.get("http://localhost:9644/v1/topics", timeout=4)
            topics = topics_resp.json() if topics_resp.status_code == 200 else []
            result["redpanda"] = {
                "status": "up",
                "is_healthy": health.get("is_healthy", True),
                "brokers": broker_count,
                "topics": len(topics),
                "topic_names": [t.get("name", t) if isinstance(t, dict) else str(t) for t in topics[:10]],
                "controller_id": health.get("controller_id", 0),
            }
        else:
            result["redpanda"] = {"status": "degraded"}
    except Exception as e:
        result["redpanda"] = {"status": "down", "error": str(e)[:80]}

    try:
        import requests as req
        ml_resp = req.get("http://localhost:8001/health", timeout=3)
        result["ml_engine"] = {"status": "up" if ml_resp.status_code == 200 else "degraded", "detail": ml_resp.json() if ml_resp.status_code == 200 else {}}
    except Exception:
        result["ml_engine"] = {"status": "down"}

    return result


class DiagnosticRequest(BaseModel):
    service: str

@app.post("/diagnostics")
def run_diagnostics(payload: DiagnosticRequest, api_key: str = Depends(verify_api_key)) -> Dict[str, Any]:
    import subprocess
    service = payload.service.lower()
    
    commands = {
        "redpanda": "docker exec cashflowscore-eventpipelineanddatalayer-redpanda-1 rpk cluster info",
        # Table is named 'events' — matches store.py schema and docker/initdb/01_schema.sql
        "postgres": "docker exec cashflowscore-eventpipelineanddatalayer-postgres-1 psql -U cashflow -d cashflowscore -c \"SELECT count(*) FROM events;\"",
        "redis": "docker exec cashflowscore-eventpipelineanddatalayer-redis-1 redis-cli info keyspace"
    }
    
    if service not in commands:
        raise HTTPException(status_code=400, detail="Invalid service for diagnostics")
        
    cmd = commands[service]
    try:
        # Run command with 10s timeout
        result = subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=10)
        
        output = result.stdout
        if result.stderr:
            output += "\n" + result.stderr
            
        return {
            "service": service,
            "command": cmd,
            "output": output.strip() or "No output returned.",
            "exit_code": result.returncode
        }
    except subprocess.TimeoutExpired:
        return {
            "service": service,
            "command": cmd,
            "output": "Command timed out after 10 seconds.",
            "exit_code": -1
        }
    except Exception as e:
        return {
            "service": service,
            "command": cmd,
            "output": f"Error executing command: {str(e)}",
            "exit_code": -1
        }

